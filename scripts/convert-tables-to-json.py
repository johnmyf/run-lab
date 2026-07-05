#!/usr/bin/env python3
"""将 data-resources 目录下的 4 个表格转换为 JSON 文件，输出到 src/static/ 供网页下载。"""

import json
import re
import os
import openpyxl

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
BASE_DIR = os.path.dirname(SCRIPT_DIR)  # 项目根目录
OUTPUT_DIR = os.path.join(BASE_DIR, "src", "static")

os.makedirs(OUTPUT_DIR, exist_ok=True)


def parse_time_to_seconds(t):
    """将 mm:ss / h:mm:ss / m:ss.s / m:ss+ 转为总秒数，解析失败返回 None"""
    if t is None or t in ("—", "——", ""):
        return None
    t = t.strip()
    # 去除末尾的 + 号 (如 "3:41+")
    if t.endswith("+"):
        t = t[:-1]
    parts = t.split(":")
    try:
        if len(parts) == 2:
            m = int(parts[0])
            s = float(parts[1])
            return m * 60 + s
        elif len(parts) == 3:
            # 处理 ":57:50" 这种缺少小时部分的情况
            if parts[0] == "":
                h = 0
            else:
                h = int(parts[0])
            m = int(parts[1])
            s = float(parts[2])
            return h * 3600 + m * 60 + s
    except (ValueError, IndexError):
        return None
    return None


def parse_md_table(filepath):
    """解析 markdown 表格，返回 (headers, rows)。"""
    with open(filepath, "r", encoding="utf-8") as f:
        lines = f.readlines()

    # 找到表格分隔行（|---|---|）之前的表头行
    table_lines = []
    in_table = False
    for line in lines:
        line = line.strip()
        if line.startswith("|") and "---" in line:
            # 分隔行，前面的就是表头
            if table_lines:
                header_line = table_lines[-1]
                break
            in_table = True
        elif line.startswith("|"):
            table_lines.append(line)
            in_table = True
        elif in_table and not line.startswith("|"):
            break

    # 找到所有的数据行（分隔行之后）
    data_lines = []
    found_sep = False
    for line in lines:
        line = line.strip()
        if line.startswith("|") and "---" in line:
            found_sep = True
            continue
        if found_sep and line.startswith("|"):
            data_lines.append(line)

    # 解析表头
    headers = [h.strip() for h in header_line.split("|")[1:-1]]
    # 解析数据行
    rows = []
    for dl in data_lines:
        cells = [c.strip() for c in dl.split("|")[1:-1]]
        if len(cells) == len(headers):
            rows.append(cells)

    return headers, rows


# ============ 表5-1: VDOT 值与比赛时间 ============
headers, rows = parse_md_table(
    os.path.join(BASE_DIR, "data-resources", "表5-1, 与常见距离的比赛时间相对应的 VDOT 值.md")
)

data_5_1 = []
for row in rows:
    vdot = int(row[0])
    entry = {"vdot": vdot}
    # row: VDOT | 1500m | 1.6k | 3k | 3.2k | 5k | 10k | 15k | 半马 | 全马 | VDOT(重复)
    distance_keys = [
        ("distance_1500m", "time_1500m"),
        ("distance_1_6km", "time_1_6km"),
        ("distance_3km", "time_3km"),
        ("distance_3_2km", "time_3_2km"),
        ("distance_5km", "time_5km"),
        ("distance_10km", "time_10km"),
        ("distance_15km", "time_15km"),
        ("distance_half_marathon", "time_half_marathon"),
        ("distance_marathon", "time_marathon"),
    ]
    # 跳过第一个 VDOT 列和最后一个重复 VDOT 列
    for i, (dist_key, time_key) in enumerate(distance_keys):
        time_str = row[i + 1]
        entry[time_key] = time_str if time_str else None
        entry[dist_key] = parse_time_to_seconds(time_str)
    data_5_1.append(entry)

with open(os.path.join(OUTPUT_DIR, "vdot-race-times.json"), "w", encoding="utf-8") as f:
    json.dump(data_5_1, f, ensure_ascii=False, indent=2)
print(f"✅ 表5-1 → vdot-race-times.json ({len(data_5_1)} 条)")

# ============ 表5-2: 基于 VDOT 的训练强度 ============
headers, rows = parse_md_table(
    os.path.join(BASE_DIR, "data-resources", "表5-2, 基于当前 VDOT 的训练强度.md")
)

data_5_2 = []
for row in rows:
    vdot = int(row[0])
    entry = {"vdot": vdot}

    # 辅助函数：解析"min~max"格式
    def parse_range(cell):
        if not cell or cell in ("—", "——"):
            return None
        parts = cell.split("~")
        if len(parts) == 2:
            return {
                "min": parts[0].strip(),
                "max": parts[1].strip(),
                "min_seconds": parse_time_to_seconds(parts[0].strip()),
                "max_seconds": parse_time_to_seconds(parts[1].strip()),
            }
        return {
            "value": cell.strip(),
            "seconds": parse_time_to_seconds(cell.strip()),
        }

    def parse_single(cell):
        """解析单个值（时间或纯数字）"""
        if not cell or cell in ("—", "——"):
            return None
        cell = cell.strip()
        seconds = parse_time_to_seconds(cell)
        if seconds is not None:
            return {"value": cell, "seconds": seconds}
        # 可能是纯数字（秒）
        try:
            return {"value": float(cell), "seconds": float(cell)}
        except ValueError:
            return {"value": cell, "seconds": None}

    # 列映射 (跳过第一个 VDOT 列)
    zone_keys = [
        ("e_l_per_km", 1),       # E/L 1km
        ("e_l_per_mile", 2),     # E/L 1.6km
        ("m_per_km", 3),         # M 1km
        ("m_per_mile", 4),       # M 1.6km
        ("t_400m", 5),           # T 400m
        ("t_per_km", 6),         # T 1km
        ("t_per_mile", 7),       # T 1.6km
        ("i_400m", 8),           # I 400m
        ("i_per_km", 9),         # I 1km
        ("i_1200m", 10),         # I 1200m
        ("i_per_mile", 11),      # I 1.6km
        ("r_200m", 12),          # R 200m
        ("r_300m", 13),          # R 300m
        ("r_400m", 14),          # R 400m
        ("r_600m", 15),          # R 600m
        ("r_800m", 16),          # R 800m
    ]

    # E/L zones have range format
    range_indices = {1, 2}

    for key, col_idx in zone_keys:
        cell = row[col_idx]
        if col_idx in range_indices:
            entry[key] = parse_range(cell)
        else:
            entry[key] = parse_single(cell)

    data_5_2.append(entry)

with open(os.path.join(OUTPUT_DIR, "vdot-training-paces.json"), "w", encoding="utf-8") as f:
    json.dump(data_5_2, f, ensure_ascii=False, indent=2)
print(f"✅ 表5-2 → vdot-training-paces.json ({len(data_5_2)} 条)")

# ============ 表5-3: 初跑者 VDOT ============
headers, rows = parse_md_table(
    os.path.join(BASE_DIR, "data-resources", "表5-3, 适用于初跑者和其他从较慢水平起步的人.md")
)

data_5_3 = []
for row in rows:
    entry = {}
    # Columns: 1.6k_time | 5k_time | VDOT | R_200m | R_300m | I_200m | I_400m | T_400m | T_1km | T_1_6km | M_time | M_per_km | M_per_1_6km
    keys = [
        "time_1_6km", "time_5km", "vdot",
        "r_200m", "r_300m", "i_200m", "i_400m",
        "t_400m", "t_per_km", "t_per_mile",
        "m_time", "m_per_km", "m_per_mile"
    ]
    for i, key in enumerate(keys):
        cell = row[i]
        if key == "vdot":
            entry[key] = int(cell) if cell.isdigit() else None
        elif cell in ("—", "——", ""):
            entry[key] = None
        else:
            seconds = parse_time_to_seconds(cell)
            entry[key] = {"value": cell, "seconds": seconds} if seconds is not None else {"value": cell, "seconds": None}
    data_5_3.append(entry)

with open(os.path.join(OUTPUT_DIR, "beginner-vdot.json"), "w", encoding="utf-8") as f:
    json.dump(data_5_3, f, ensure_ascii=False, indent=2)
print(f"✅ 表5-3 → beginner-vdot.json ({len(data_5_3)} 条)")

# ============ 表15-2: 休整后训练量调整 ============
wb = openpyxl.load_workbook(
    os.path.join(BASE_DIR, "data-resources", "表15-2 休整后训练量的调整.xlsx")
)
ws = wb["休整后训练量调整表"]

# 读全部行
all_rows = []
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
    all_rows.append([cell if cell is not None else "" for cell in row])

# 数据结构：categories
# 第1行是标题
# 第2行是表头：类别 | 休整时间 | 对训练量的调整 | 休整前的VDOT（%）
# 从第3行开始是数据
categories = []
current_category = None

for row in all_rows[2:]:  # 跳过标题行和表头行
    cat_num = str(row[0]).strip() if row[0] else ""
    rest_time = str(row[1]).strip() if row[1] else ""
    adjustment = str(row[2]).strip() if row[2] else ""
    vdot_pct = str(row[3]).strip() if row[3] else ""

    # 跳过空行
    if not cat_num and not rest_time and not adjustment and not vdot_pct:
        continue
    # 跳过注释行
    if cat_num.startswith("注"):
        continue

    if cat_num in ("1", "2", "3", "4"):
        current_category = {
            "category": int(cat_num),
            "rest_period": rest_time,
            "adjustments": [{"adjustment": adjustment, "vdot_percent": vdot_pct}],
        }
        categories.append(current_category)
    elif current_category is not None:
        current_category["adjustments"].append(
            {"adjustment": adjustment, "vdot_percent": vdot_pct, "rest_time": rest_time}
        )

with open(os.path.join(OUTPUT_DIR, "training-volume-adjustment.json"), "w", encoding="utf-8") as f:
    json.dump(categories, f, ensure_ascii=False, indent=2)
print(f"✅ 表15-2 → training-volume-adjustment.json ({len(categories)} 个类别)")


print(f"\n所有 JSON 文件已生成到: {OUTPUT_DIR}")
print("文件列表:")
for f in sorted(os.listdir(OUTPUT_DIR)):
    if f.endswith(".json"):
        size = os.path.getsize(os.path.join(OUTPUT_DIR, f))
        print(f"  - {f} ({size:,} bytes)")
